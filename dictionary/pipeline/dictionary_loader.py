"""
Dictionary Loader — Parses the governance dictionary Excel file.
Extracts Object metadata, Fields definitions, and Lists values.
"""
import openpyxl


def load_dictionary(path: str) -> dict:
    """Load and parse all three sheets from the governance dictionary."""
    wb = openpyxl.load_workbook(path, data_only=True)

    obj = _parse_object_sheet(wb["Object"])
    fields = _parse_fields_sheet(wb["Fields"])

    wb.close()
    return {"object": obj, "fields": fields}


def _parse_object_sheet(ws) -> dict:
    """Parse the Object sheet: row 1 = headers, row 2 = values."""
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    obj = dict(zip(headers, values))

    # Normalize types
    if obj.get("expected_column_count") is not None:
        obj["expected_column_count"] = int(obj["expected_column_count"])
    if obj.get("invalid_rows_tolerance_pct") is not None:
        obj["invalid_rows_tolerance_pct"] = int(obj["invalid_rows_tolerance_pct"])

    return obj


def _parse_fields_sheet(ws) -> list[dict]:
    """Parse the Fields sheet: row 1 = headers, rows 2+ = field definitions."""
    headers = [cell.value for cell in ws[1]]
    fields = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        field = dict(zip(headers, row))

        # Normalize
        if field.get("length") is not None:
            field["length"] = int(field["length"])
        if field.get("decimal") is not None:
            field["decimal"] = int(field["decimal"])
        if field.get("integers") is not None:
            field["integers"] = int(field["integers"])

        # Parse allowed_values into a list
        av = field.get("allowed_values")
        if av and isinstance(av, str):
            field["allowed_values"] = [v.strip() for v in av.split(",")]
        else:
            field["allowed_values"] = None

        fields.append(field)

    return fields
